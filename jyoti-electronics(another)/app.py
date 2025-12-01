import os
from datetime import datetime, timedelta, date
from io import StringIO, BytesIO
import csv
import logging

from flask import Flask, render_template, request, redirect, url_for, send_from_directory, flash, jsonify, Response, send_file
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func

# CONFIG
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DB_PATH = os.path.join(BASE_DIR, 'jyoti_electronics.db')
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')  # (unused but kept)
SCREENSHOT_PATH = '/mnt/data/e1931f15-4839-40fb-8711-d4c3d283d76b.jpeg'  # your uploaded image

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app = Flask(__name__)
app.config['SECRET_KEY'] = 'change-me'
app.config['SQLALCHEMY_DATABASE_URI'] = f"sqlite:///{DB_PATH}"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# optional: path to wkhtmltopdf binary (set in environment if needed)
# e.g. export WKHTMLTOPDF_PATH="/usr/local/bin/wkhtmltopdf"
WKHTMLTOPDF_PATH = os.environ.get('WKHTMLTOPDF_PATH')

# configure simple logging
logging.basicConfig(level=logging.INFO)
app.logger.setLevel(logging.INFO)

db = SQLAlchemy(app)

# MODELS
class Customer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    phone = db.Column(db.String(50))
    address = db.Column(db.String(500))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    jobs = db.relationship('Job', backref='customer', cascade="all, delete-orphan")

class Job(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=False)

    # spreadsheet-like fields
    area = db.Column(db.String(200))
    tv_model = db.Column(db.String(300))
    repair_work = db.Column(db.Text)
    amount_charged = db.Column(db.Float, default=0.0)
    expense = db.Column(db.Float, default=0.0)  # job-level expense
    payment_mode = db.Column(db.String(50))
    pickup_date = db.Column(db.String(50))
    note = db.Column(db.Text)
    status = db.Column(db.String(50), default='received')  # received, in_progress, completed

    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    completed_at = db.Column(db.DateTime, nullable=True)

    payments = db.relationship('Payment', backref='job', cascade="all, delete-orphan")

    @property
    def profit(self):
        return (self.amount_charged or 0.0) - (self.expense or 0.0)

class Payment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    job_id = db.Column(db.Integer, db.ForeignKey('job.id'), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    payment_mode = db.Column(db.String(50))
    note = db.Column(db.String(200))
    payment_date = db.Column(db.DateTime, default=datetime.utcnow)

# --- NEW: Expense model ---
class Expense(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    description = db.Column(db.String(300))
    amount = db.Column(db.Float, nullable=False, default=0.0)
    date = db.Column(db.Date, default=date.today)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# ensure tables created (existing + new)
with app.app_context():
    db.create_all()

# HELPERS
def job_total_paid(job):
    return sum(p.amount for p in job.payments) if job.payments else 0.0

# ROUTES (your original routes left intact)
@app.route('/')
def index():
    # show simple form-first page
    recent_jobs = Job.query.order_by(Job.created_at.desc()).limit(10).all()
    return render_template('index.html', recent_jobs=recent_jobs, screenshot_url=SCREENSHOT_PATH)

@app.route('/jobs')
def jobs():
    all_jobs = Job.query.order_by(Job.created_at.desc()).all()
    return render_template('jobs.html', jobs=all_jobs, job_total_paid=job_total_paid)

@app.route('/job/<int:job_id>')
def job_detail(job_id):
    j = Job.query.get_or_404(job_id)
    total_paid = job_total_paid(j)
    remaining = (j.amount_charged or 0.0) - total_paid
    return render_template('job_detail.html', job=j, total_paid=total_paid, remaining=remaining)

@app.route('/customer/new', methods=['POST'])
def new_customer():
    name = request.form.get('name')
    if not name:
        flash('Name required', 'danger')
        return redirect(url_for('index'))
    c = Customer(name=name, phone=request.form.get('phone'), address=request.form.get('address'))
    db.session.add(c)
    db.session.commit()
    flash('Customer created', 'success')
    return redirect(url_for('index'))

@app.route('/new_job', methods=['POST'])
def new_job():
    # allow selecting existing customer by id or creating inline
    cid = request.form.get('customer_id')
    if cid:
        customer = Customer.query.get(int(cid))
    else:
        cname = request.form.get('cust_name')
        if not cname:
            flash('Customer name required', 'danger')
            return redirect(url_for('index'))
        customer = Customer(name=cname, phone=request.form.get('cust_phone'), address=request.form.get('cust_address'))
        db.session.add(customer)
        db.session.commit()

    amount = float(request.form.get('amount_charged') or 0)
    expense = float(request.form.get('expense') or 0)
    job = Job(
        customer_id=customer.id,
        area=request.form.get('area'),
        tv_model=request.form.get('tv_model'),
        repair_work=request.form.get('repair_work'),
        amount_charged=amount,
        expense=expense,
        payment_mode=request.form.get('payment_mode'),
        pickup_date=request.form.get('pickup_date'),
        note=request.form.get('note'),
        status='received'
    )
    db.session.add(job)
    db.session.commit()

    # if advance provided, store as payment
    adv = float(request.form.get('advance_amount') or 0)
    if adv > 0:
        p = Payment(job_id=job.id, amount=adv, payment_mode=request.form.get('payment_mode'), note='advance')
        db.session.add(p); db.session.commit()

    flash('Job created', 'success')
    return redirect(url_for('job_detail', job_id=job.id))

@app.route('/job/<int:job_id>/add_payment', methods=['POST'])
def add_payment(job_id):
    job = Job.query.get_or_404(job_id)
    amt = float(request.form.get('amount') or 0)
    if amt <= 0:
        flash('Amount must be positive', 'danger')
        return redirect(request.referrer or url_for('job_detail', job_id=job_id))
    p = Payment(job_id=job.id, amount=amt, payment_mode=request.form.get('payment_mode'), note=request.form.get('note'))
    db.session.add(p); db.session.commit()
    flash('Payment recorded', 'success')
    return redirect(request.referrer or url_for('job_detail', job_id=job_id))

@app.route('/job/<int:job_id>/complete', methods=['POST'])
def complete(job_id):
    job = Job.query.get_or_404(job_id)
    job.status = 'completed'
    job.completed_at = datetime.utcnow()
    db.session.commit()
    flash('Job marked completed', 'success')
    return redirect(request.referrer or url_for('job_detail', job_id=job_id))

@app.route('/job/<int:job_id>/delete', methods=['POST'])
def delete_job(job_id):
    job = Job.query.get_or_404(job_id)
    db.session.delete(job)
    db.session.commit()
    flash('Job deleted', 'success')
    return redirect(url_for('jobs'))

# ------------------ EXPENSE ROUTES (robust, includes alias) ------------------

@app.route('/expenses')
@app.route('/expenses/')
@app.route('/daily_expenses')
@app.route('/daily_expenses/')
def expenses_list():
    try:
        start = request.args.get('start')
        end = request.args.get('end')

        q = Expense.query

        if start:
            try:
                s = datetime.strptime(start, "%Y-%m-%d").date()
                q = q.filter(Expense.date >= s)
            except Exception:
                app.logger.warning("Invalid start date filter: %s", start)

        if end:
            try:
                e = datetime.strptime(end, "%Y-%m-%d").date()
                q = q.filter(Expense.date <= e)
            except Exception:
                app.logger.warning("Invalid end date filter: %s", end)

        items = q.order_by(Expense.date.desc(), Expense.created_at.desc()).all()

        # compute total on server side
        total_amount = sum((float(it.amount) if it.amount is not None else 0.0) for it in items)

        # Debug log: list amounts seen (remove after debug)
        app.logger.info("Expenses count=%d, total_amount=%s", len(items), total_amount)
        for it in items:
            app.logger.info("  expense id=%s date=%s desc=%s amount=%s", it.id, getattr(it, 'date', None), it.description, it.amount)

        return render_template('expenses.html', expenses=items, total_amount=total_amount)

    except Exception as exc:
        app.logger.exception("Error rendering expenses_list")
        return f"Server error while loading expenses page: {exc}", 500


@app.route('/expenses/new', methods=['GET', 'POST'])
def new_expense():
    if request.method == 'POST':
        desc = request.form.get('description')
        try:
            amt = float(request.form.get('amount') or 0)
        except Exception:
            flash('Invalid amount', 'danger')
            return redirect(url_for('expenses_list'))
        d = request.form.get('date')
        try:
            date_parsed = datetime.strptime(d, "%Y-%m-%d").date() if d else datetime.utcnow().date()
        except Exception:
            date_parsed = datetime.utcnow().date()

        if amt <= 0:
            flash("Amount must be positive", "danger")
            return redirect(url_for('expenses_list'))

        e = Expense(description=desc, amount=amt, date=date_parsed)
        db.session.add(e)
        db.session.commit()
        flash("Expense added", "success")
        return redirect(url_for('expenses_list'))

    # GET
    return render_template("new_expense.html")

@app.route('/expenses/<int:expense_id>/delete', methods=['POST'])
def delete_expense(expense_id):
    e = Expense.query.get_or_404(expense_id)
    db.session.delete(e)
    db.session.commit()
    flash("Expense deleted", "success")
    return redirect(url_for('expenses_list'))

# ------------------ INVOICE ROUTES ------------------

@app.route('/job/<int:job_id>/invoice')
def invoice_html(job_id):
    j = Job.query.get_or_404(job_id)
    total_paid = job_total_paid(j)
    remaining = (j.amount_charged or 0) - total_paid
    return render_template(
        "invoice.html",
        job=j,
        total_paid=total_paid,
        remaining=remaining,
        generated_at=datetime.utcnow()
    )

@app.route('/job/<int:job_id>/invoice/pdf')
def invoice_pdf(job_id):
    j = Job.query.get_or_404(job_id)
    total_paid = job_total_paid(j)
    html = render_template(
        "invoice.html",
        job=j,
        total_paid=total_paid,
        remaining=(j.amount_charged or 0) - total_paid,
        generated_at=datetime.utcnow()
    )

    try:
        # try pdfkit (wkhtmltopdf). If WKHTMLTOPDF_PATH is set, use it.
        import pdfkit
        config = None
        if WKHTMLTOPDF_PATH:
            try:
                config = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_PATH)
            except Exception as cex:
                app.logger.warning("Could not configure pdfkit with WKHTMLTOPDF_PATH=%s: %s", WKHTMLTOPDF_PATH, cex)
                config = None
        pdf = pdfkit.from_string(html, False, configuration=config)
        return Response(pdf, mimetype='application/pdf', headers={"Content-Disposition": f"attachment;filename=invoice_{job_id}.pdf"})
    except Exception as e:
        app.logger.exception("PDF generation failed for job %s", job_id)
        # fallback: return the HTML invoice; flash a visible message
        flash("PDF not available. Use Print → Save as PDF.", "warning")
        return html

# daily summary API (original)
@app.route('/api/daily_summary')
def daily_summary():
    days = int(request.args.get('days', 7))
    today = datetime.utcnow().date()
    start = today - timedelta(days=days-1)

    pay_q = (db.session.query(func.date(Payment.payment_date).label('day'),
                              func.sum(Payment.amount).label('payments_total'),
                              func.count(Payment.id).label('payments_count'))
             .filter(func.date(Payment.payment_date) >= start)
             .group_by(func.date(Payment.payment_date)).all())

    job_q = (db.session.query(func.date(Job.completed_at).label('day'),
                              func.sum(Job.amount_charged).label('jobs_total'),
                              func.sum(Job.expense).label('jobs_expense'),
                              func.count(Job.id).label('jobs_count'))
             .filter(Job.completed_at != None)
             .filter(func.date(Job.completed_at) >= start)
             .group_by(func.date(Job.completed_at)).all())

    summary = {}
    for i in range(days):
        d = (start + timedelta(days=i)).isoformat()
        summary[d] = {'date': d, 'payments_total': 0.0, 'payments_count': 0, 'jobs_amount_total': 0.0, 'jobs_expense_total': 0.0, 'jobs_profit_total': 0.0, 'jobs_completed_count': 0}

    for r in pay_q:
        day = r.day.isoformat() if isinstance(r.day, (datetime, date)) else str(r.day)
        summary[day]['payments_total'] = float(r.payments_total or 0.0)
        summary[day]['payments_count'] = int(r.payments_count or 0)

    for r in job_q:
        day = r.day.isoformat() if isinstance(r.day, (datetime, date)) else str(r.day)
        summary[day]['jobs_amount_total'] = float(r.jobs_total or 0.0)
        summary[day]['jobs_expense_total'] = float(r.jobs_expense or 0.0)
        summary[day]['jobs_profit_total'] = float((r.jobs_total or 0.0) - (r.jobs_expense or 0.0))
        summary[day]['jobs_completed_count'] = int(r.jobs_count or 0)

    out = [summary[d] for d in sorted(summary.keys())]
    totals = {'range_payments_total': sum(x['payments_total'] for x in out), 'range_jobs_amount_total': sum(x['jobs_amount_total'] for x in out), 'range_jobs_expense_total': sum(x['jobs_expense_total'] for x in out), 'range_jobs_profit_total': sum(x['jobs_profit_total'] for x in out)}
    return jsonify({'summary': out, 'totals': totals})

# ---- NEW: daily summary including expenses (non-destructive new endpoint) ----
@app.route('/api/daily_summary_with_expenses')
def daily_summary_with_expenses():
    days = int(request.args.get('days', 7))
    today = datetime.utcnow().date()
    start = today - timedelta(days=days-1)

    pay_q = (db.session.query(func.date(Payment.payment_date).label('day'),
                              func.sum(Payment.amount).label('payments_total'),
                              func.count(Payment.id).label('payments_count'))
             .filter(func.date(Payment.payment_date) >= start)
             .group_by(func.date(Payment.payment_date)).all())

    job_q = (db.session.query(func.date(Job.completed_at).label('day'),
                              func.sum(Job.amount_charged).label('jobs_total'),
                              func.sum(Job.expense).label('jobs_expense'),
                              func.count(Job.id).label('jobs_count'))
             .filter(Job.completed_at != None)
             .filter(func.date(Job.completed_at) >= start)
             .group_by(func.date(Job.completed_at)).all())

    expense_q = (db.session.query(func.date(Expense.date).label('day'),
                                  func.sum(Expense.amount).label('expenses_total'),
                                  func.count(Expense.id).label('expenses_count'))
                 .filter(func.date(Expense.date) >= start)
                 .group_by(func.date(Expense.date)).all())

    summary = {}
    for i in range(days):
        d = (start + timedelta(days=i)).isoformat()
        summary[d] = {'date': d, 'payments_total': 0.0, 'payments_count': 0, 'jobs_amount_total': 0.0, 'jobs_expense_total': 0.0, 'jobs_profit_total': 0.0, 'jobs_completed_count': 0, 'expenses_total': 0.0, 'expenses_count': 0, 'net_profit_total': 0.0}

    for r in pay_q:
        day = r.day.isoformat() if hasattr(r.day, 'isoformat') else str(r.day)
        summary[day]['payments_total'] = float(r.payments_total or 0.0)
        summary[day]['payments_count'] = int(r.payments_count or 0)
    for r in job_q:
        day = r.day.isoformat() if hasattr(r.day, 'isoformat') else str(r.day)
        summary[day]['jobs_amount_total'] = float(r.jobs_total or 0.0)
        summary[day]['jobs_expense_total'] = float(r.jobs_expense or 0.0)
        summary[day]['jobs_profit_total'] = float((r.jobs_total or 0.0) - (r.jobs_expense or 0.0))
        summary[day]['jobs_completed_count'] = int(r.jobs_count or 0)
    for r in expense_q:
        day = r.day.isoformat() if hasattr(r.day, 'isoformat') else str(r.day)
        summary[day]['expenses_total'] = float(r.expenses_total or 0.0)
        summary[day]['expenses_count'] = int(r.expenses_count or 0)
    for k, v in summary.items():
        v['net_profit_total'] = v['jobs_profit_total'] - v['expenses_total']

    out = [summary[d] for d in sorted(summary.keys())]
    totals = {
        'range_payments_total': sum(x['payments_total'] for x in out),
        'range_jobs_amount_total': sum(x['jobs_amount_total'] for x in out),
        'range_jobs_expense_total': sum(x['jobs_expense_total'] for x in out),
        'range_jobs_profit_total': sum(x['jobs_profit_total'] for x in out),
        'range_expenses_total': sum(x['expenses_total'] for x in out),
        'range_net_profit_total': sum(x['net_profit_total'] for x in out)
    }
    return jsonify({'summary': out, 'totals': totals})

@app.route('/daily_summary.csv')
def daily_summary_csv():
    res = daily_summary().get_json() if isinstance(daily_summary(), Response) else daily_summary()
    data = res if isinstance(res, dict) else res
    out = StringIO()
    w = csv.writer(out)
    w.writerow(['Date','Payments','#payments','Jobs amount','Jobs expense','Profit','#jobs'])
    for r in data['summary']:
        w.writerow([r['date'], f"{r['payments_total']:.2f}", r['payments_count'], f"{r['jobs_amount_total']:.2f}", f"{r['jobs_expense_total']:.2f}", f"{r['jobs_profit_total']:.2f}", r['jobs_completed_count']])
    return Response(out.getvalue(), mimetype='text/csv', headers={'Content-Disposition':'attachment;filename=daily_summary.csv'})

@app.route('/export_jobs.csv')
def export_jobs_csv():
    out = StringIO(); w = csv.writer(out)
    w.writerow(['JobID','Date','Customer','Phone','Area','TVModel','RepairWork','Charge','Expense','Profit','PaymentMode','Note','Status','Pickup','CompletedAt'])
    for j in Job.query.order_by(Job.created_at.desc()).all():
        w.writerow([j.id, j.created_at.strftime('%Y-%m-%d %H:%M'), j.customer.name, j.customer.phone or '', j.area or '', j.tv_model or '', j.repair_work or '', f"{j.amount_charged:.2f}", f"{j.expense:.2f}", f"{j.profit:.2f}", j.payment_mode or '', j.note or '', j.status or '', j.pickup_date or '', j.completed_at.strftime('%Y-%m-%d %H:%M') if j.completed_at else ''])
    return Response(out.getvalue(), mimetype='text/csv', headers={'Content-Disposition':'attachment;filename=jobs_export.csv'})

# optional xlsx
try:
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    @app.route('/export_jobs.xlsx')
    def export_jobs_xlsx():
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Jobs"
        headers = ['JobID','Date','Customer','Phone','Area','TVModel','RepairWork','Charge','Expense','Profit','PaymentMode','Note','Status','Pickup','CompletedAt']
        ws.append(headers)
        for cell in ws[1]: cell.font = Font(bold=True)
        for j in Job.query.order_by(Job.created_at.desc()).all():
            ws.append([j.id, j.created_at.strftime('%Y-%m-%d %H:%M'), j.customer.name, j.customer.phone or '', j.area or '', j.tv_model or '', j.repair_work or '', float(j.amount_charged), float(j.expense), float(j.profit), j.payment_mode or '', j.note or '', j.status or '', j.pickup_date or '', j.completed_at.strftime('%Y-%m-%d %H:%M') if j.completed_at else ''])
        for col in ws.columns:
            max_len=0; col_letter=get_column_letter(col[0].column)
            for cell in col:
                v=str(cell.value or '')
                if len(v)>max_len: max_len=len(v)
            ws.column_dimensions[col_letter].width = min(50, max_len+2)
        bio = BytesIO(); wb.save(bio); bio.seek(0)
        return send_file(bio, as_attachment=True, download_name='jobs.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
except Exception:
    pass

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
